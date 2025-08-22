"""
Módulo para exportar la tabla pivote a Excel, con formato especial:
- Cada departamento en una carpeta
- Cada municipio en una hoja
- Nombre del municipio en B2, fecha en G2, tabla desde B4
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime


def exportar_pivot_a_excel(pivot_pob_full, output_dir):
    departamentos = pivot_pob_full.columns.get_level_values("NOM_DTO_DPN").unique()

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for depto in departamentos:
        if pd.isna(depto):
            continue

        depto_folder = os.path.join(output_dir, str(depto))
        if not os.path.exists(depto_folder):
            os.makedirs(depto_folder)

        excel_path = os.path.join(depto_folder, "No. 1. Población obj.xlsx")

        municipios = pivot_pob_full.columns.get_level_values("NOM_M_DPN")[
            pivot_pob_full.columns.get_level_values("NOM_DTO_DPN") == depto
        ].unique()

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            for municipio in municipios:
                if pd.isna(municipio):
                    continue
                try:
                    df_slice = pivot_pob_full.xs(
                        (depto, municipio), axis=1, level=("NOM_DTO_DPN", "NOM_M_DPN")
                    )
                except KeyError:
                    continue
                df_slice.to_excel(
                    writer,
                    sheet_name=str(municipio)[:31],
                    startrow=3,
                    startcol=1,
                    index=True,
                    header=True,
                )

        wb = load_workbook(excel_path)
        for municipio in municipios:
            if pd.isna(municipio):
                continue
            ws = wb[str(municipio)[:31]]
            ws["B2"] = "Municipio:"
            ws["B2"].font = Font(bold=True, color="000000")
            ws["C2"] = str(municipio)
            ws["C2"].font = Font(bold=True, color="000000")
            ws["F2"] = "Fecha de Actualización:"
            ws["F2"].font = Font(bold=True, color="000000")
            ws["G2"] = datetime.now().strftime("%Y-%m-%d")
            ws["G2"].font = Font(bold=True, color="000000")

            # Texto y formato en B15, B16, B17
            ws["B15"] = "Número estimado de mujeres con intención reproductiva:"
            ws["B15"].font = Font(bold=True, color="000000")
            ws["B16"] = "Número estimado de gestantes:"
            ws["B16"].font = Font(bold=True, color="000000")
            ws["B17"] = "Total de nacidos vivos:"
            ws["B17"].font = Font(bold=True, color="000000")

            # Aplica bordes a B15, B16, B17
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ["B15", "B16", "B17"]:
                ws[cell].border = border

        wb.save(excel_path)
        print(f"Creado: {excel_path}")
